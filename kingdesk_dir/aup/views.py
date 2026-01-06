from django.shortcuts import render, redirect
from aup.models import Employers, Posts, Divisions, Current_shedules, Months, Genders, Future_shedules, Week_days, Current_tasks, Future_tasks, Future_shedules_New_Year, Wishes
from django.views.generic import DetailView
from datetime import datetime, timedelta
from django.contrib.auth.models import User
import requests
import openpyxl
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill

def error(request): 
    hbr_employers = ['h-01', 'h-02', 'h-03', 'h-04', 'h-05']
    aup_employers = ['a-01', 'a-02', 'a-03']
    employers = Employers.objects.order_by('surname')
    user = None
    id_user = ''
    for i in employers:
        if i.username == request.user.username:
            user = i
            break
    
    return render(request, 'sign_in/error_page.html', {'user':user, 'employers':employers, 'hbr_employers':hbr_employers, 'aup_employers':aup_employers, 'id_user':id_user})

def current_shedule(request): 
    date_now = datetime.now().date()
    hbr_employers = ['h-01', 'h-02', 'h-03', 'h-04', 'h-05']
    aup_employers = ['a-01', 'a-02', 'a-03']
    employers = Employers.objects.order_by('surname')
    current_shedules = Current_shedules.objects.all()
    months = Months.objects.all()
    user = None
    id_user = ''
    for i in employers:
        if i.username == request.user.username:
            user = i
            break
    return render(request, 'aup/current_shedule.html', {'user':user, 'employers':employers, 'hbr_employers':hbr_employers, 'aup_employers':aup_employers, 'current_shedules':current_shedules, 'months':months, 'id_user':id_user, 'date_now':date_now})

def wishes(request): 
    employers = Employers.objects.order_by('surname')
    hbr_employers = ['h-01', 'h-02', 'h-03', 'h-04', 'h-05']
    aup_employers = ['a-01', 'a-02', 'a-03']
    wishes = Wishes.objects.all()
    user = None
    id_user = ''
    for i in employers:
        if i.username == request.user.username:
            user = i
            break
    return render(request, 'aup/wishes.html', {'user':user, 'employers':employers, 'hbr_employers':hbr_employers, 'aup_employers':aup_employers, 'id_user':id_user, 'wishes':wishes})

def send_wish(request, user_id):
    wishes = Wishes.objects.all()
    employers = Employers.objects.all()
    for i in employers:
        if i.id == user_id:
            wish = Wishes()
            wish.employer_from = i.surname + ' ' + i.name
            wish.wish = request.POST.get('message')
            wish.employer_to = request.POST.get('employer_to')
            wish.save()
            print()
            break
    
    return redirect ('aup_wishes') 

def future_shedule(request):
    hbr_employers = ['h-01', 'h-02', 'h-03', 'h-04', 'h-05']
    aup_employers = ['a-01', 'a-02', 'a-03']
    employers = Employers.objects.order_by('surname')
    future_shedules = Future_shedules.objects.all()
    months = Months.objects.all()
    user = None
    id_user = ''
    for i in employers:
        if i.username == request.user.username:
            user = i
            break
    return render(request, 'aup/future_shedule.html', {'user':user, 'employers':employers, 'hbr_employers':hbr_employers, 'aup_employers':aup_employers, 'future_shedules':future_shedules, 'months':months, 'id_user':id_user})

def future_shedule_hny(request):
    hbr_employers = ['h-01', 'h-02', 'h-03', 'h-04', 'h-05']
    aup_employers = ['a-01', 'a-02', 'a-03']
    employers = Employers.objects.order_by('surname')
    future_shedules = Future_shedules_New_Year.objects.all()
    months = Months.objects.all()
    user = None
    id_user = ''
    for i in employers:
        if i.username == request.user.username:
            user = i
            break
    return render(request, 'aup/future_shedule_hny.html', {'user':user, 'employers':employers, 'hbr_employers':hbr_employers, 'aup_employers':aup_employers, 'future_shedules':future_shedules, 'months':months, 'id_user':id_user})

def edit_current_shedule(request, user_id):
    current_shedules = Current_shedules.objects.all()
    for i in current_shedules:
        if request.POST.get("date") == str(i.date.day) + ' ' + str(i.date.month) and user_id == i.employer_id.id:
            id = i.id
            break
    shedule = Current_shedules.objects.get(id=id)
    try:
        if request.method == "POST":
            if request.POST.get("is_weekend") == 'on':
                shedule.is_weekend = True
                shedule.start_work = '00:00:00'
                shedule.end_work = '00:00:00'
            else:
                shedule.is_weekend = False
                shedule.start_work = request.POST.get("start_work")
                shedule.end_work = request.POST.get("end_work")
            shedule.save()
        return redirect ('aup_cs')
    except:
        return redirect ('aup_cs')      

def edit_future_shedule(request, user_id):
    future_shedules = Future_shedules.objects.all()
    for i in future_shedules:
        if request.POST.get("date") == str(i.date.day) + ' ' + str(i.date.month) and user_id == i.employer_id.id:
            id = i.id
            break
    shedule = Future_shedules.objects.get(id=id)
    try:
        if request.method == "POST":
            if request.POST.get("is_weekend") == 'on':
                shedule.is_weekend = True
                shedule.start_work = '00:00:00'
                shedule.end_work = '00:00:00'
            else:
                shedule.is_weekend = False
                shedule.start_work = request.POST.get("start_work")
                shedule.end_work = request.POST.get("end_work")
            shedule.save()
        return redirect ('aup_fs')
    except:
        return redirect ('aup_fs')

def edit_future_shedule_hny(request, user_id):
    future_shedules = Future_shedules_New_Year.objects.all()
    for i in future_shedules:
        if request.POST.get("date") == str(i.date.day) + ' ' + str(i.date.month) and user_id == i.employer_id.id:
            id = i.id
            break
    shedule = Future_shedules_New_Year.objects.get(id=id)
    try:
        if request.method == "POST":
            if request.POST.get("is_weekend") == 'on':
                shedule.is_weekend = True
                shedule.start_work = '00:00:00'
                shedule.end_work = '00:00:00'
            else:
                shedule.is_weekend = False
                shedule.start_work = request.POST.get("start_work")
                shedule.end_work = request.POST.get("end_work")
            shedule.save()
        return redirect ('aup_fs_hny')
    except:
        return redirect ('aup_fs_hny')

def profile(request):
    hbr_employers = ['h-01', 'h-02', 'h-03', 'h-04', 'h-05']
    aup_employers = ['a-01', 'a-02', 'a-03']
    employers = Employers.objects.order_by('surname')
    user = None
    for i in employers:
        if i.username == request.user.username:
            user = i
            break
    return render(request, 'aup/profile.html', {'user':user, 'employers':employers, 'hbr_employers':hbr_employers, 'aup_employers':aup_employers})

def edit_profile(request):
    hbr_employers = ['h-01', 'h-02', 'h-03', 'h-04', 'h-05']
    aup_employers = ['a-01', 'a-02', 'a-03']
    employers = Employers.objects.order_by('surname')
    posts = Posts.objects.order_by('post_name')
    months = Months.objects.all()
    divisions = Divisions.objects.order_by('division_name')
    user = None
    for i in employers:
        if i.username == request.user.username:
            user = i
            break
    return render(request, 'aup/edit_profile.html', {'user':user, 'employers':employers, 'hbr_employers':hbr_employers, 'aup_employers':aup_employers, 'posts':posts, 'months':months, 'divisions':divisions})

def update_profile(request, user_id):
    try:
        if request.method == "POST":
            user = Employers.objects.get(id=user_id)

            user.surname = request.POST.get("surname")
            user.name = request.POST.get("name")
            user.lastname = request.POST.get("lastname")
            user.phone = request.POST.get("phone")
            user.birthday = request.POST.get("birthday")

            if user.password != request.POST.get("password"):
                user.password = request.POST.get("password")
                user.save()
                user_2 = User.objects.get(username = user.username)
                user_2.delete() 
                User.objects.create_user(username=user.username, password=user.password)
            else:
                user.save()

            if user.post_id.post_code != 'a-01':
                post = Posts.objects.get(id=request.POST.get("post"))
                division = Divisions.objects.get(id=request.POST.get("division"))
                user.division_id = division
                if user.post_id != post:
                    user.post_id = post
                    user.save()
                    return redirect('logout')
            user.save()
        return redirect ('aup_profile')
    except:
        return redirect ('aup_edit_profile')
    
def employers(request):
    date_now = datetime.now().date()
    hbr_employers = ['h-01', 'h-02', 'h-03', 'h-04', 'h-05']
    aup_employers = ['a-01', 'a-02', 'a-03']
    employers = Employers.objects.order_by('surname')
    user = None
    for i in employers:
        if i.username == request.user.username:
            user = i
            break
    return render(request, 'aup/employers.html', {'user':user, 'date_now':date_now, 'employers':employers, 'hbr_employers':hbr_employers, 'aup_employers':aup_employers})

class Employer_Detail_View(DetailView):
    model = Employers
    template_name = 'aup/employer_info.html'
    context_object_name = 'employer'

def add_employer(request):
    hbr_employers = ['h-01', 'h-02', 'h-03', 'h-04', 'h-05']
    aup_employers = ['a-01', 'a-02', 'a-03']
    employers = Employers.objects.order_by('surname')
    posts = Posts.objects.order_by('post_name')
    genders = Genders.objects.all()
    divisions = Divisions.objects.order_by('division_name')
    user = None
    
    for i in employers:
        if i.username == request.user.username:
            user = i
            break
    return render(request, 'aup/add_employer.html', {'user':user, 'employers':employers, 'hbr_employers':hbr_employers, 'aup_employers':aup_employers, 'posts':posts, 'divisions':divisions, 'genders':genders})

def register_employer (request):
    hbr_employers = ['h-01', 'h-02', 'h-03', 'h-04', 'h-05']
    try:
        if request.method == "POST":
            user = Employers()
            user.surname = request.POST.get("surname")
            user.name = request.POST.get("name")
            user.lastname = request.POST.get("lastname")
            user.username = request.POST.get("username")
            user.phone = request.POST.get("phone")
            post = Posts.objects.get(id=request.POST.get("post"))
            division = Divisions.objects.get(id=request.POST.get("division"))
            gender = Genders.objects.get(id=request.POST.get("gender"))
            user.gender_id = gender
            user.division_id = division
            user.post_id = post
            user.photo = request.FILES.get("photo")
            user.password = request.POST.get("password")
            user.birthday = request.POST.get("birthday")
            User.objects.create_user(username=request.POST.get("username"), password=request.POST.get("password"), email=request.POST.get("email"))
            user.save()

            employers = Employers.objects.all()
            current_shedules = Current_shedules.objects.all()
            future_shedules = Future_shedules.objects.all()
            count = 0
            for i in employers:
                if count < 1:
                    if i.post_id.post_code in hbr_employers:
                        count += 1
                        for j in current_shedules:
                            if j.employer_id.id == i.id:
                                current_shedule = Current_shedules()
                                current_shedule.week_day_id = Week_days.objects.get(id=j.week_day_id.id)
                                current_shedule.date = j.date
                                current_shedule.employer_id = Employers.objects.get(username = request.POST.get("username"))
                                current_shedule.start_work = '00:00:00'
                                current_shedule.end_work = '00:00:00'
                                current_shedule.is_weekend = True
                                current_shedule.save()
                        for j in future_shedules:
                            if j.employer_id.username == i.username:
                                future_shedule = Future_shedules()
                                future_shedule.week_day_id = Week_days.objects.get(id=j.week_day_id.id)
                                future_shedule.date = j.date
                                future_shedule.employer_id = Employers.objects.get(username = request.POST.get("username"))
                                future_shedule.start_work = '00:00:00'
                                future_shedule.end_work = '00:00:00'
                                future_shedule.is_weekend = True
                                future_shedule.comment = ''
                                future_shedule.save()
                else:
                    break
                
        return redirect ('aup_employers')
    except:
        return redirect ('aup_add_employer')

def delete_employer (request, user_id):
    user = Employers.objects.get(id=user_id)
    username = user.username
    user_2 = User.objects.get(username = username)
    user_2.delete() 
    user.delete()
    return redirect ('aup_employers')
    
def edit_employer(request, user_id):
    hbr_employers = ['h-01', 'h-02', 'h-03', 'h-04', 'h-05']
    aup_employers = ['a-01', 'a-02', 'a-03']
    employers = Employers.objects.order_by('surname')
    posts = Posts.objects.order_by('post_name')
    genders = Genders.objects.all()
    months = Months.objects.all()
    divisions = Divisions.objects.order_by('division_name')
    user = None
    employer = Employers.objects.get(id=user_id)
    
    for i in employers:
        if i.username == request.user.username:
            user = i
            break
    return render(request, 'aup/edit_employer.html', {'user':user, 'employers':employers, 'months':months, 'hbr_employers':hbr_employers, 'aup_employers':aup_employers, 'posts':posts, 'divisions':divisions, 'genders':genders, 'employer':employer})

def update_employer(request, user_id):
    try:
        if request.method == "POST":
            employer = Employers.objects.get(id=user_id)
            employers = Employers.objects.order_by('surname')
            user = None
            for i in employers:
                if i.username == request.user.username:
                    user = i
                    break
            employer.surname = request.POST.get("surname")
            employer.name = request.POST.get("name")
            employer.lastname = request.POST.get("lastname")
            employer.phone = request.POST.get("phone")
            employer.birthday = request.POST.get("birthday")

            if employer.password != request.POST.get("password"):
                employer.password = request.POST.get("password")
                employer.save()
                employer_2 = User.objects.get(username = employer.username)
                employer_2.delete() 
                User.objects.create_user(username=employer.username, password=employer.password)
            else:
                employer.save()

            if employer.username != user.username:
                post = Posts.objects.get(id=request.POST.get("post"))
                division = Divisions.objects.get(id=request.POST.get("division"))
                employer.division_id = division
                employer.post_id = post
            else:
                if employer.post_id.post_code != 'a-01':
                    post = Posts.objects.get(id=request.POST.get("post"))
                    division = Divisions.objects.get(id=request.POST.get("division"))
                    employer.division_id = division
                    employer.post_id = post
            employer.save()
        return redirect ('aup_employers')
    except:
        return redirect ('aup_employers')

def stop_future(request, division_id):
    division = Divisions.objects.get(id=division_id)
    if division.is_stop == True:
        division.is_stop = False
    else:
        division.is_stop = True
    division.save()
    return redirect ('aup_fs')

def update_shedule(request, division_id):
    current_shedules = Current_shedules.objects.all() 
    future_shedules = Future_shedules.objects.all()
    division = Divisions.objects.get(id=division_id)
    for i in current_shedules:
        if i.employer_id.division_id.id == division.id:
            current_shedule = Current_shedules.objects.get(id=i.id)
            current_shedule.delete()
    for i in future_shedules:
        if i.employer_id.division_id.id == division.id:
            current_shedule = Current_shedules()
            current_shedule.week_day_id = Week_days.objects.get(id=i.week_day_id.id)
            current_shedule.date = i.date
            current_shedule.employer_id = Employers.objects.get(id=i.employer_id.id)
            current_shedule.start_work = i.start_work
            current_shedule.end_work = i.end_work
            current_shedule.is_weekend = i.is_weekend
            current_shedule.save()
    current_shedules = Current_shedules.objects.all()
    for i in future_shedules:
        if i.employer_id.division_id.id == division.id:
            future_shedule = Future_shedules.objects.get(id=i.id)
            future_shedule.delete()
    future_shedules = Future_shedules.objects.all()
    for i in current_shedules:
        if i.employer_id.division_id.id == division_id:
            future_shedule = Future_shedules()
            future_shedule.week_day_id = Week_days.objects.get(id=i.week_day_id.id)
            future_shedule.date = i.date + timedelta(days=7)
            future_shedule.employer_id = Employers.objects.get(id=i.employer_id.id)
            future_shedule.start_work = '00:00:00'
            future_shedule.end_work = '00:00:00'
            future_shedule.is_weekend = True
            future_shedule.comment = ''
            future_shedule.save()
    division.is_stop = False
    division.save()
    return redirect ('aup_cs')

def current_tasks(request): 
    hbr_employers = ['h-01', 'h-02', 'h-03', 'h-04', 'h-05']
    aup_employers = ['a-01', 'a-02', 'a-03']
    employers = Employers.objects.order_by('surname')
    current_shedules = Current_shedules.objects.all()
    current_tasks = Current_tasks.objects.all()
    months = Months.objects.all()
    user = None
    id_user = ''
    for i in employers:
        if i.username == request.user.username:
            user = i
            break
    return render(request, 'aup/current_tasks.html', {'user':user, 'employers':employers, 'hbr_employers':hbr_employers, 'aup_employers':aup_employers, 'current_shedules':current_shedules, 'months':months, 'id_user':id_user, 'current_tasks':current_tasks})
                  
def future_tasks(request): 
    hbr_employers = ['h-01', 'h-02', 'h-03', 'h-04', 'h-05']
    aup_employers = ['a-01', 'a-02', 'a-03']
    employers = Employers.objects.order_by('surname')
    current_shedules = Current_shedules.objects.all()
    months = Months.objects.all()
    user = None
    id_user = ''
    for i in employers:
        if i.username == request.user.username:
            user = i
            break
    return render(request, 'aup/future_tasks.html', {'user':user, 'employers':employers, 'hbr_employers':hbr_employers, 'aup_employers':aup_employers, 'current_shedules':current_shedules, 'months':months, 'id_user':id_user})

def reload(request, division_id): 
    current_shedules = Current_shedules.objects.all()
    future_shedules = Future_shedules.objects.all()
    division = Divisions.objects.get(id=division_id)

    for i in current_shedules:
        if i.employer_id.division_id.id == division.id:
            current_shedule = Current_shedules.objects.get(id=i.id)
            current_shedule.delete()
    for i in future_shedules:
        if i.employer_id.division_id.id == division.id:
            current_shedule = Current_shedules()
            current_shedule.week_day_id = Week_days.objects.get(id=i.week_day_id.id)
            current_shedule.date = i.date - timedelta(7)
            current_shedule.employer_id = Employers.objects.get(id=i.employer_id.id)
            current_shedule.start_work = i.start_work
            current_shedule.end_work = i.end_work
            current_shedule.is_weekend = i.is_weekend
            current_shedule.save()
            future_shedule = Future_shedules.objects.get(id=i.id)
            future_shedule.delete()
    for i in current_shedules:
        if i.employer_id.division_id.id == division_id:
            future_shedule = Future_shedules()
            future_shedule.week_day_id = Week_days.objects.get(id=i.week_day_id.id)
            future_shedule.date = i.date + timedelta(7)
            future_shedule.employer_id = Employers.objects.get(id=i.employer_id.id)
            future_shedule.start_work = '00:00:00'
            future_shedule.end_work = '00:00:00'
            future_shedule.is_weekend = True
            future_shedule.comment = ''
            future_shedule.save()
    return redirect ('aup_cs')

def export_to_excel(request):
    try:
        web_dates=[]
        dates=[]
        table_start =[]
        table_end = []
        employers_table = []
        employers = Employers.objects.order_by('surname')
        current_shedules = Current_shedules.objects.all()
        user = None
        for i in employers:
            if i.username == request.user.username:
                user = i
                break
        
        wb = load_workbook(f'./downloads/Raspisanie_{user.division_id.division_name[3:]}.xlsx')
        sheet = wb.active
        index_dates = 0
        for row in sheet:
            if index_dates == 1:
                for column in row:
                    if column.value!=None:
                        dates.append(column)
                break
            else:
                index_dates+=1
        for i in current_shedules:
            if i.employer_id.id == user.id:
                web_dates.append(i.date)
        index_dates = 0
        for i in dates:
            sheet.cell(row=i.row, column=i.column).value = web_dates[index_dates]
            index_dates+=1
        for row in sheet:
            for i in employers:
                if i.surname + ' ' + i.name == row[0].value:
                    a = []
                    b=[]
                    index = 0
                    for value in row:
                        if index>=4:
                            if index%2 == 0:
                                a.append(value)
                            else:
                                b.append(value)
                        if index == 0:
                            employers_table.append(value)  
                        index+=1
                    table_start.append(a)
                    table_end.append(b)
        for i in current_shedules:
            index = 0
            for j in employers_table:
                jndex = 0
                if j.value == i.employer_id.surname + ' ' + i.employer_id.name:
                    for k in dates:
                        if str(k.value) == str(i.date):
                            shedule = Current_shedules.objects.get(id=i.id)
                            if shedule.is_weekend == False:
                                sheet.cell(row=table_start[index][jndex].row, column=table_start[index][jndex].column).value = shedule.start_work.strftime('%H:%M')
                                sheet.cell(row=table_end[index][jndex].row, column=table_end[index][jndex].column).value = shedule.end_work.strftime('%H:%M')
                                if shedule.start_work == user.division_id.start_work:
                                    sheet.cell(row=table_start[index][jndex].row, column=table_start[index][jndex].column).value = '08:30'
                                    sheet.cell(row=table_start[index][jndex].row, column=table_start[index][jndex].column).fill = PatternFill(fill_type='solid', fgColor='95b3d7')
                                    sheet.cell(row=table_end[index][jndex].row, column=table_end[index][jndex].column).fill = PatternFill(fill_type='solid', fgColor='95b3d7')
                                elif shedule.end_work == user.division_id.end_work:
                                    sheet.cell(row=table_start[index][jndex].row, column=table_start[index][jndex].column).fill = PatternFill(fill_type='solid', fgColor='b1a0c7')
                                    sheet.cell(row=table_end[index][jndex].row, column=table_end[index][jndex].column).fill = PatternFill(fill_type='solid', fgColor='b1a0c7')
                                else:
                                    sheet.cell(row=table_start[index][jndex].row, column=table_start[index][jndex].column).fill = PatternFill(fill_type='solid', fgColor='ffffff')
                                    sheet.cell(row=table_end[index][jndex].row, column=table_end[index][jndex].column).fill = PatternFill(fill_type='solid', fgColor='ffffff')
                            else:
                                sheet.cell(row=table_start[index][jndex].row, column=table_start[index][jndex].column).value = ''
                                sheet.cell(row=table_end[index][jndex].row, column=table_end[index][jndex].column).value = ''
                                sheet.cell(row=table_start[index][jndex].row, column=table_start[index][jndex].column).fill = PatternFill(fill_type='solid', fgColor='ffff99')
                                sheet.cell(row=table_end[index][jndex].row, column=table_end[index][jndex].column).fill = PatternFill(fill_type='solid', fgColor='ffff99')
                        jndex+=1
                index+=1
        wb.save(f'./downloads/Raspisanie_{user.division_id.division_name[3:]}.xlsx')
        
        return redirect('aup_cs')
    except:
        return redirect('error')

def export_to_excel_future(request):
    try:
        web_dates=[]
        dates=[]
        table_start =[]
        table_end = []
        employers_table = []
        employers = Employers.objects.order_by('surname')
        future_shedules = Future_shedules.objects.all()
        user = None
        for i in employers:
            if i.username == request.user.username:
                user = i
                break
        
        wb = load_workbook(f'./downloads/Raspisanie_{user.division_id.division_name[3:]}.xlsx')
        sheet = wb.active
        index_dates = 0
        for row in sheet:
            if index_dates == 1:
                for column in row:
                    if column.value!=None:
                        dates.append(column)
                break
            else:
                index_dates+=1
        for i in future_shedules:
            if i.employer_id.id == user.id:
                web_dates.append(i.date)
        index_dates = 0
        for i in dates:
            sheet.cell(row=i.row, column=i.column).value = web_dates[index_dates]
            index_dates+=1
        for row in sheet:
            for i in employers:
                if i.surname + ' ' + i.name == row[0].value:
                    a = []
                    b=[]
                    index = 0
                    for value in row:
                        if index>=4:
                            if index%2 == 0:
                                a.append(value)
                            else:
                                b.append(value)
                        if index == 0:
                            employers_table.append(value)  
                        index+=1
                    table_start.append(a)
                    table_end.append(b)
        for i in future_shedules:
            index = 0
            for j in employers_table:
                jndex = 0
                if j.value == i.employer_id.surname + ' ' + i.employer_id.name:
                    for k in dates:
                        if str(k.value) == str(i.date):
                            shedule = Future_shedules.objects.get(id=i.id)
                            if shedule.is_weekend == False:
                                sheet.cell(row=table_start[index][jndex].row, column=table_start[index][jndex].column).value = shedule.start_work.strftime('%H:%M')
                                sheet.cell(row=table_end[index][jndex].row, column=table_end[index][jndex].column).value = shedule.end_work.strftime('%H:%M')
                                if shedule.start_work == user.division_id.start_work:
                                    sheet.cell(row=table_start[index][jndex].row, column=table_start[index][jndex].column).value = '08:30'
                                    sheet.cell(row=table_start[index][jndex].row, column=table_start[index][jndex].column).fill = PatternFill(fill_type='solid', fgColor='95b3d7')
                                    sheet.cell(row=table_end[index][jndex].row, column=table_end[index][jndex].column).fill = PatternFill(fill_type='solid', fgColor='95b3d7')
                                elif shedule.end_work == user.division_id.end_work:
                                    sheet.cell(row=table_start[index][jndex].row, column=table_start[index][jndex].column).fill = PatternFill(fill_type='solid', fgColor='b1a0c7')
                                    sheet.cell(row=table_end[index][jndex].row, column=table_end[index][jndex].column).fill = PatternFill(fill_type='solid', fgColor='b1a0c7')
                                else:
                                    sheet.cell(row=table_start[index][jndex].row, column=table_start[index][jndex].column).fill = PatternFill(fill_type='solid', fgColor='ffffff')
                                    sheet.cell(row=table_end[index][jndex].row, column=table_end[index][jndex].column).fill = PatternFill(fill_type='solid', fgColor='ffffff')
                            else:
                                sheet.cell(row=table_start[index][jndex].row, column=table_start[index][jndex].column).value = ''
                                sheet.cell(row=table_end[index][jndex].row, column=table_end[index][jndex].column).value = ''
                                sheet.cell(row=table_start[index][jndex].row, column=table_start[index][jndex].column).fill = PatternFill(fill_type='solid', fgColor='ffff99')
                                sheet.cell(row=table_end[index][jndex].row, column=table_end[index][jndex].column).fill = PatternFill(fill_type='solid', fgColor='ffff99')
                        jndex+=1
                index+=1
        wb.save(f'./downloads/Raspisanie_{user.division_id.division_name[3:]}.xlsx')
        
        return redirect('aup_fs')
    except:
        return redirect('error')

def import_from_excel(request):
    try:
        employers = Employers.objects.order_by('surname')
        current_shedules = Current_shedules.objects.all()
        user = None
        for i in employers:
            if i.username == request.user.username:
                user = i
                break
        
        wb = load_workbook(f'./downloads/Raspisanie_{user.division_id.division_name[3:]}.xlsx')
        sheet = wb.active
        table_start =[]
        table_end = []
        dates = []
        index_dates = 0
        for row in sheet.values:
            if index_dates == 1:
                for column in row:
                    if column!=None:
                        dates.append(column)
                break
            else:
                index_dates+=1
        employers_table = []
        for row in sheet.values:
            for i in employers:
                if i.surname + ' ' + i.name == row[0]:
                    a = []
                    b=[]
                    index = 0
                    for value in row:
                        if index>=4:
                            if value != None and value!='Выходной':
                                if type(value) == str:
                                    time = datetime.strptime(value, '%H:%M')
                                    if index%2 == 0:
                                        if time.strftime('%H:%M') == '08:30':
                                            a.append('08:00')
                                        else:
                                            a.append(time.strftime('%H:%M'))
                                    else:
                                        b.append(time.strftime('%H:%M'))
                                else:
                                    if index%2 == 0:
                                        if value.strftime('%H:%M') == '08:30':
                                            a.append('08:00')
                                        else:
                                            a.append(value.strftime('%H:%M'))
                                    else:
                                        b.append(value.strftime('%H:%M'))
                            else:
                                if index%2==0:
                                    a.append('Выходной') 
                                else:
                                    b.append('Выходной')
                        if index == 0:
                            employers_table.append(value)  
                        index+=1
                    table_start.append(a)
                    table_end.append(b)
                    
        for i in current_shedules:
            index = 0
            for j in employers_table:
                jndex = 0
                if j == i.employer_id.surname + ' ' + i.employer_id.name:
                    for k in dates:
                        if str(k.date()) == str(i.date):
                            shedule = Current_shedules.objects.get(id=i.id)
                            if table_start[index][jndex] != None and table_start[index][jndex] != 'Выходной':
                                shedule.start_work = table_start[index][jndex]
                                shedule.end_work = table_end[index][jndex]
                                shedule.is_weekend = False
                            elif table_start[index][jndex] == None or table_start[index][jndex] == 'Выходной':
                                shedule.start_work = '00:00:00'
                                shedule.end_work = '00:00:00'
                                shedule.is_weekend = True
                            else:
                                shedule.start_work = '00:00:00'
                                shedule.end_work = '00:00:00'
                                shedule.is_weekend = True
                            shedule.save()
                        jndex+=1
                index+=1
        return redirect('aup_cs')
    except:
        return redirect('error')

    